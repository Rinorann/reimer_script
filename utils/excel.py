from openpyxl import load_workbook

def load_full_plate(file_path, sheet_name):
    '''
    принимает на вход путь к файлу с синерджи и название листа в нём.
    выдаёт словарь типа 'A1': od для всего планшета 8x12.
    '''
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    # 1. Ищем "Результаты" в колонке A
    results_row = None
    for row in range(1, 150):
        if ws.cell(row=row, column=1).value == "Результаты":
            results_row = row
            break
    
    if results_row is None:
        raise ValueError("Не найдено слово 'Результаты' в колонке A")
    
    # 2. A1 планшета = строка results_row + 4, колонка 3 (C)
    a1_row = results_row + 4
    a1_col = 3  # C = 3
    
    plate_data = {}
    rows = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    
    for i, row_letter in zip(range(0, 32, 5), rows): # целится по read 2:550 по каждой букве в планшете
        for col in range(12):  # 12 колонок планшета
            well = f"{row_letter}{col+1}"
            od_value = ws.cell(
                row=a1_row + i,           # A -> +0, B -> +1, C -> +2...
                column=a1_col + col       # 1я колонка -> +0, 2я -> +1...
            ).value
            plate_data[well] = od_value
    
    wb.close()
    return plate_data


