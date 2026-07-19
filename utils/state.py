# utils/state.py
from openpyxl import load_workbook

class State:
    def __init__(self):
        self.wells = {}  # dict well: OD (из Excel)
        # Структура пирамиды: { sample_name: { "ferritin_mass": float, "dilutions": { dil_str: [wells] } } }
        self.samples = {}  
        self.calibration_points = []  # list of {"concentration": float, "wells": list}
        self.regression_coef = None  # {"k": k, "b": b, "R2": R2, "center_od": center_od}

    def load_plate(self, file_path, sheet_name='Sheet1'):
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        results_row = None
        for row in range(1, 150):
            if ws.cell(row=row, column=1).value == "Результаты":
                results_row = row
                break
        if results_row is None:
            raise ValueError("Не найдено слово 'Результаты' в колонке A")
        
        a1_row = results_row + 4
        a1_col = 3
        plate_data = {}
        rows = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        for i, row_letter in zip(range(0, 32, 5), rows):
            for col in range(12):
                well = f"{row_letter}{col+1}"
                od_value = ws.cell(row=a1_row + i, column=a1_col + col).value
                plate_data[well] = od_value
        wb.close()
        self.wells = plate_data

    def _get_od_of_well(self, wells: list):
        ods = []
        for well in wells:
            if well in self.wells and isinstance(self.wells[well], (float, int)):
                ods.append(self.wells[well])
        return ods

    def calculate_calibration(self):
        """Строит регрессию по точкам калибровки: OD = k * Концентрация + b"""
        if not self.calibration_points: return None
        x_vals, y_vals = [], []

        for pt in self.calibration_points:
            ods = self._get_od_of_well(pt["wells"])
            if ods:
                x_vals.append(pt["concentration"])
                y_vals.append(sum(ods) / len(ods))

        n = len(x_vals)
        if n < 2: return None

        sum_x = sum(x_vals)
        sum_y = sum(y_vals)
        sum_xx = sum(x * x for x in x_vals)
        sum_xy = sum(x * y for x, y in zip(x_vals, y_vals))

        denominator = (n * sum_xx - sum_x * sum_x)
        if denominator == 0: return None

        k = (n * sum_xy - sum_x * sum_y) / denominator
        b = (sum_y - k * sum_x) / n

        mean_y = sum_y / n
        ss_tot = sum((y - mean_y) ** 2 for y in y_vals)
        ss_res = sum((y - (k * x + b)) ** 2 for x, y in zip(x_vals, y_vals))
        r_squared = 1 - (ss_res / ss_tot) if ss_tot != 0 else 1.0

        # Математический центр калибровочной кривой
        center_od = sum(y_vals) / len(y_vals)
        self.regression_coef = {"k": k, "b": b, "R2": r_squared, "center_od": center_od}
        return self.regression_coef

    def calculate_samples(self):
        """Алгоритм пирамиды: Выбор оптимального разведения по центру кривой и расчет атомов Fe"""
        self.calculate_calibration()
        if not self.regression_coef: return {}

        k = self.regression_coef["k"]
        b = self.regression_coef["b"]
        center_od = self.regression_coef["center_od"]

        M_Fe = 55.845
        M_Ferritin = 450000.0
        
        results = {}

        for sample_name, sample_data in self.samples.items():
            ferritin_mass_mkg = sample_data["ferritin_mass"]
            dils_dict = sample_data["dilutions"]

            best_dil = None
            best_raw_od = None
            min_diff = float('inf')
            best_wells = []

            # 1. Поиск разведения, среднее значение которого ближе всего к центру калибровки
            for dil_str, wells in dils_dict.items():
                ods = self._get_od_of_well(wells)
                if not ods: continue
                
                mean_raw_od = sum(ods) / len(ods)
                diff = abs(mean_raw_od - center_od)
                if diff < min_diff:
                    min_diff = diff
                    best_dil = float(dil_str)
                    best_raw_od = mean_raw_od
                    best_wells = wells

            if best_dil is None: continue

            # 2. Пересчет OD и концентрации железа на исходный неразведенный раствор
            final_od = best_raw_od * best_dil
            fe_conc_origin = (final_od - b) / k if k != 0 else 0.0

            # 3. Подсчет атомов Fe на одну молекулу белка ферритина
            if ferritin_mass_mkg > 0 and fe_conc_origin > 0:
                moles_fe = fe_conc_origin / M_Fe
                moles_ferritin = ferritin_mass_mkg / M_Ferritin
                atoms_per_ferritin = round(moles_fe / moles_ferritin, 1)
            else:
                atoms_per_ferritin = "N/A"

            results[sample_name] = {
                "ferritin_mass": ferritin_mass_mkg,
                "chosen_dil": best_dil,
                "chosen_raw_od": best_raw_od,
                "wells": best_wells,
                "final_od": final_od,
                "fe_conc": fe_conc_origin,
                "atoms_fe": atoms_per_ferritin,
                "all_dils": list(dils_dict.keys())
            }
            
        return results